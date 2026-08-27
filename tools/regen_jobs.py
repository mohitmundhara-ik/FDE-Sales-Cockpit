#!/usr/bin/env python3
"""
regen_jobs.py - rebuild the cockpit's JOBS dataset from a raw LinkedIn scrape.

Run this whenever a fresh scrape lands. It reproduces the tagging the original
dataset used, so the cockpit's numbers stay internally consistent rather than
drifting between passes.

    python3 regen_jobs.py <scrape.xlsx> --out JOBS.json

Sheets expected: one named for FDE-titled roles, one named 'Adjacent'.
Everything is derived from descriptionText, so no manual tagging is required.
"""
import sys, json, re, argparse, unicodedata
from collections import defaultdict
import openpyxl

# ---------- the eight blocks, in the cockpit's own order and weights ----------
BLOCKS = [
    ("prodeng",   "Production engineering", 29),
    ("custcraft", "Customer craft",         20),
    ("aieng",     "AI engineering",         16),
    ("soldesign", "Solution design",        13),
    ("deployops", "Deployment & operations", 8),
    ("dataint",   "Data & integration",      6),
    ("evalq",     "Evaluation & quality",    5),
    ("secgov",    "Security & governance",   3),
]

# ---------- stack vocabulary, grouped by the block it evidences ----------
STACK = {
 "prodeng": [
   ("Python", r"\bpython\b"), ("Java", r"\bjava\b(?!script)"),
   ("JavaScript / TypeScript", r"\b(javascript|typescript|node\.?js)\b"),
   ("Go", r"\bgolang\b|\bgo\b(?= |,|/|\))"), ("C++ / C#", r"\bc\+\+|\bc#"),
   ("REST / APIs", r"\brest\b|\bapis?\b|\bgraphql\b"), ("GraphQL", r"\bgraphql\b"),
   ("Microservices", r"microservice"), ("React / frontend", r"\breact\b|\bangular\b|\bvue\b|frontend"),
   ("Git", r"\bgit\b|github|gitlab"), ("FastAPI / Flask / Django", r"fastapi|flask|django"),
 ],
 "custcraft": [
   ("Stakeholder management", r"stakeholder"), ("Executive communication", r"executive|c-level|c-suite|leadership team"),
   ("POCs / pilots", r"\bpoc\b|proof of concept|\bpilot"), ("Discovery & workshops", r"discovery|workshop"),
   ("Presales / demos", r"pre-?sales|\bdemo"), ("Onsite / embedded", r"on-?site|embedded with|customer site|client site"),
   ("Requirements gathering", r"requirement"), ("Training & enablement", r"enablement|train(ing)? (the )?(customer|client|user)"),
 ],
 "aieng": [
   ("LLM APIs", r"\bllms?\b|large language model|openai|gpt-?\d"),
   ("RAG / retrieval", r"\brag\b|retrieval[- ]augmented"),
   ("Agents / agentic", r"\bagentic\b|\bai agents?\b|multi-?agent"),
   ("LangChain / LangGraph", r"langchain|langgraph|llamaindex"),
   ("Anthropic / Claude", r"anthropic|\bclaude\b"),
   ("Vector databases", r"vector (db|database|store)|pinecone|weaviate|qdrant|chroma|faiss"),
   ("Prompt engineering", r"prompt engineer|prompting"),
   ("Fine-tuning", r"fine-?tun|\blora\b|\bpeft\b"),
   ("MCP", r"\bmcp\b|model context protocol"),
 ],
 "soldesign": [
   ("System design", r"system design|architect(ure|ing)?\b"), ("Trade-off analysis", r"trade-?offs?"),
   ("Scalability", r"scalab|\bat scale\b"), ("Solution architecture", r"solution architect"),
   ("Technical roadmap", r"roadmap"),
 ],
 "deployops": [
   ("Cloud (AWS/GCP/Azure)", r"\baws\b|\bgcp\b|google cloud|\bazure\b"),
   ("Kubernetes / Docker", r"kubernetes|\bk8s\b|docker|container"),
   ("CI/CD", r"ci/?cd|continuous (integration|deployment)|jenkins"),
   ("Terraform / IaC", r"terraform|infrastructure as code|\biac\b|ansible|pulumi"),
   ("Observability", r"observab|monitoring|datadog|prometheus|grafana|logging"),
   ("On-call / incident", r"on-?call|incident|\bsre\b|reliability"),
   ("Serverless", r"serverless|lambda|cloud function"),
 ],
 "dataint": [
   ("SQL", r"\bsql\b"), ("ETL / pipelines", r"\betl\b|data pipeline|airflow|\bdbt\b"),
   ("Kafka / streaming", r"kafka|streaming|pub/?sub"),
   ("Warehouses", r"snowflake|databricks|bigquery|redshift|warehouse"),
   ("Data modelling", r"data model|schema design"),
 ],
 "evalq": [
   ("Evals / benchmarking", r"\bevals?\b|evaluation|benchmark"),
   ("Testing", r"\btest(ing|s)?\b|\bqa\b|pytest|unit test"),
   ("Guardrails", r"guardrail|hallucinat|safety"),
   ("Quality metrics", r"accuracy|precision|recall|\bf1\b"),
 ],
 "secgov": [
   ("Security", r"\bsecurity\b|\bsoc ?2\b|encryption|vulnerab"),
   ("Compliance", r"compliance|\bgdpr\b|\bhipaa\b|\bdpdp\b|regulat"),
   ("Auth / RBAC", r"\brbac\b|\bsso\b|\boauth\b|\bsaml\b|authentication|authoriz"),
   ("Audit / governance", r"audit|governance|data residency"),
 ],
}

# a block counts as "asked for" when the posting names at least one of its signals
BLOCK_RX = {k: re.compile("|".join(p for _, p in v), re.I) for k, v in STACK.items()}
STACK_RX = {k: [(n, re.compile(p, re.I)) for n, p in v] for k, v in STACK.items()}

# ---------- level ----------
LEVEL_RULES = [
    ("Head / Director", r"\b(head of|director|vp\b|vice president|chief)\b"),
    ("Manager",         r"\b(manager|managing)\b"),
    ("Lead / Principal", r"\b(principal|staff|lead\b|architect|distinguished|senior manager)\b"),
    ("Senior",          r"\b(senior|sr\.?|iii\b)\b"),
    ("Associate",       r"\b(associate|junior|jr\.?|entry|graduate|intern|\bi\b)\b"),
]
SENIORITY_MAP = {
    "Internship": "Associate", "Entry level": "Associate", "Associate": "Associate",
    "Mid-Senior level": "Mid", "Director": "Head / Director", "Executive": "Head / Director",
}

def level_of(title, seniority, years=None):
    """Reverse-engineered from the original dataset and kept identical so level
    counts stay comparable between passes: a seniority word in the title wins,
    and when there is none the years stated in the posting decide. In the
    original, Associate only ever held 0-2 years or none, and Mid only 3-5."""
    t = title or ""
    for lv, rx in LEVEL_RULES:
        if re.search(rx, t, re.I):
            return lv
    if years is None or years <= 2: return "Associate"
    return "Mid"          # the original never promoted past Mid without a title word

# ---------- employer tier ----------
SERVICES = r"(accenture|infosys|wipro|tcs|tata consultancy|cognizant|capgemini|hcl|tech mahindra|ltimindtree|mphasis|coforge|persistent|globallogic|epam|zensar|birlasoft|mindtree|deloitte|pwc|\bey\b|kpmg|mckinsey|bain|bcg|genpact|concentrix|nagarro|ust\b|virtusa|hexaware|cyient|quest global|itc infotech|happiest minds|sonata|zoho corp)"
AI_STARTUP = r"(openai|anthropic|sarvam|krutrim|cohere|scale ai|glean|sierra|decagon|harvey|cognition|perplexity|mistral|fiddler|truefoundry|cartesia|livekit|elevenlabs|qdrant|weaviate|pinecone|langchain|skit\.?ai|avaamo|neuron7|whatfix|uniphase|observe\.?ai|yellow\.?ai|haptik)"
PRODUCT = r"(google|microsoft|amazon|aws|meta|apple|netflix|salesforce|adobe|oracle|sap\b|ibm|nvidia|databricks|snowflake|servicenow|atlassian|mongodb|cloudflare|palo alto|vmware|intuit|workday|zoom|slack|stripe|paypal|uber|airbnb|linkedin|twilio|datadog|splunk|elastic|hubspot|zendesk|freshworks|zoho|razorpay|phonepe|paytm|swiggy|zomato|flipkart|myntra|meesho|groww|cred|dream11|nykaa|lenskart|urban company|browserstack|postman|chargebee|druva|icertis|innovaccer|sprinklr|gupshup|mindtickle|highradius|darwinbox|clevertap|moengage|netcore|capillary|vymo|exotel|knowlarity|kaleyra|sinch|infobip|route mobile|tanla|nice\b|verint|genesys|avaya|five9|talkdesk|8x8|ringcentral|dialpad|aircall|zoominfo|outreach|salesloft|gong|chorus|clari|people\.?ai)"

def tier_of(company, industries, employees):
    c = (company or "").lower()
    if re.search(SERVICES, c): return "Services / GCC / consulting"
    if re.search(AI_STARTUP, c): return "AI startup / frontier"
    if re.search(PRODUCT, c): return "Product company"
    try:
        n = int(float(employees or 0))
    except (TypeError, ValueError):
        n = 0
    ind = (industries or "").lower()
    if n and n < 500 and ("software" in ind or "technology" in ind or "artificial" in ind):
        return "AI startup / frontier"
    if re.search(r"(consult|outsourc|staffing|it services)", ind):
        return "Services / GCC / consulting"
    return "Other / unclassified"

# ---------- city ----------
CITY_RULES = [
    ("Bengaluru", r"bengaluru|bangalore"), ("Hyderabad", r"hyderabad|telangana|secunderabad"),
    ("Delhi NCR", r"delhi|gurugram|gurgaon|noida|ncr\b|faridabad|ghaziabad"),
    ("Mumbai", r"mumbai|bombay|navi mumbai|thane"), ("Pune", r"\bpune\b|pimpri"),
    ("Chennai", r"chennai|madras"), ("Kolkata", r"kolkata|calcutta"),
    ("Ahmedabad", r"ahmedabad|gandhinagar"), ("Coimbatore", r"coimbatore"),
    ("Kochi", r"kochi|cochin|ernakulam|trivandrum|thiruvananthapuram"),
    ("Jaipur", r"jaipur"), ("Indore", r"indore"), ("Chandigarh", r"chandigarh|mohali"),
    ("Bhubaneswar", r"bhubaneswar"), ("Nagpur", r"nagpur"), ("Vadodara", r"vadodara|baroda"),
    ("Visakhapatnam", r"visakhapatnam|vizag"), ("Mysuru", r"mysuru|mysore"),
    ("Remote / India-wide", r"\bremote\b|work from home|anywhere"),
]
INDIA_RX = re.compile("|".join(p for _, p in CITY_RULES) + r"|india|karnataka|maharashtra|tamil nadu|telangana|haryana|uttar pradesh|west bengal|gujarat|kerala|rajasthan|punjab|odisha|madhya pradesh|andhra", re.I)
def in_india(loc):
    """Location strings like 'Greater Bengaluru Area' never say India, so match on
    the city and state vocabulary as well as the country name."""
    return bool(INDIA_RX.search(loc or ""))

def city_of(loc):
    l = (loc or "")
    for name, rx in CITY_RULES:
        if re.search(rx, l, re.I): return name
    return "India (unspecified)"

# ---------- stated years ----------
YEARS_RX = re.compile(r"(\d{1,2})\s*\+?\s*(?:-|to|–)?\s*(\d{1,2})?\s*(?:\+)?\s*years?", re.I)
def years_of(text):
    if not text: return None
    m = YEARS_RX.search(text[:4000])
    if not m: return None
    try:
        y = int(m.group(1))
        return y if 0 < y <= 25 else None
    except ValueError:
        return None

def clean(s):
    if s is None: return ""
    return unicodedata.normalize("NFKC", str(s)).strip()

def read_sheet(ws, band):
    hdr = [clean(c.value) for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdr) if h}
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        g = lambda k: row[idx[k]] if k in idx and idx[k] < len(row) else None
        title = clean(g("title"))
        if not title: continue
        loc = clean(g("location"))
        if loc and not in_india(loc):
            continue                      # India scope only
        desc = clean(g("descriptionText"))
        yrs = years_of(desc)
        posted = g("postedAt")
        out.append({
            "id": clean(g("id")) or clean(g("link")),
            "t": title, "c": clean(g("companyName")) or "Unknown",
            "city": city_of(loc), "lv": level_of(title, clean(g("seniorityLevel")), yrs),
            "b": band, "y": yrs,
            "p": (str(posted)[:10] if posted else ""),
            "u": clean(g("link")),
            "ct": tier_of(clean(g("companyName")), clean(g("industries")), g("companyEmployeesCount")),
            "_d": desc,
        })
    return out

def tag(job):
    d = job.pop("_d", "")
    hay = (job["t"] + " " + d)
    job["bl"] = [k for k in BLOCK_RX if BLOCK_RX[k].search(hay)]
    s = []
    for k, pairs in STACK_RX.items():
        for name, rx in pairs:
            if rx.search(hay) and name not in s:
                s.append(name)
    job["s"] = s
    return job

def aggregate(jobs, band):
    rows = [j for j in jobs if j["b"] == band]
    levels = ["all"] + [lv for _, lv in LEVEL_RULES][::-1]
    levels = ["all", "Associate", "Mid", "Senior", "Lead / Principal", "Manager", "Head / Director"]
    agg = {}
    for lv in levels:
        sub = rows if lv == "all" else [j for j in rows if j["lv"] == lv]
        if not sub:
            continue
        n = len(sub)
        blocks = {k: round(100 * sum(1 for j in sub if k in j["bl"]) / n) for k, _, _ in BLOCKS}
        stack = {}
        for k, pairs in STACK_RX.items():
            counts = []
            for name, _rx in pairs:
                c = sum(1 for j in sub if name in j["s"])
                if c:
                    counts.append([name, round(100 * c / n)])
            counts.sort(key=lambda x: -x[1])
            stack[k] = counts
        agg[lv] = {"n": n, "blocks": blocks, "stack": stack}
    return agg

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--out", default="JOBS.json")
    ap.add_argument("--core-sheet", default=None)
    a = ap.parse_args()

    wb = openpyxl.load_workbook(a.xlsx, read_only=False, data_only=True)
    names = wb.sheetnames
    core_name = a.core_sheet or next(n for n in names if re.search(r"fde|forward", n, re.I))
    adj_name = next(n for n in names if re.search(r"adjacent", n, re.I))

    raw_core = read_sheet(wb[core_name], "Core FDE")
    raw_adj = read_sheet(wb[adj_name], "Adjacent")
    raw = {"Core FDE": len(raw_core), "Adjacent": len(raw_adj)}

    # de-duplicate within band on posting id, then across bands (a role can appear in both)
    seen, jobs = set(), []
    for j in raw_core + raw_adj:
        key = j["id"] or (j["t"] + "|" + j["c"] + "|" + j["city"] + "|" + j["p"])
        if key in seen: continue
        seen.add(key)
        jobs.append(tag(j))

    uniq = {b: sum(1 for j in jobs if j["b"] == b) for b in ("Core FDE", "Adjacent")}
    # a role re-posted three times is one opening, not three - report it alongside
    roles = {b: len({(j["t"].lower().strip(), j["c"].lower().strip(), j["city"])
                     for j in jobs if j["b"] == b}) for b in ("Core FDE", "Adjacent")}

    out = {
        "counts": {"raw": raw, "uniq": uniq, "roles": roles},
        "blocks": [[k, n, w] for k, n, w in BLOCKS],
        "agg": {b: aggregate(jobs, b) for b in ("Core FDE", "Adjacent")},
        "levels": {b: {lv: sum(1 for j in jobs if j["b"] == b and j["lv"] == lv)
                       for lv in ["Associate", "Mid", "Senior", "Lead / Principal", "Manager", "Head / Director"]}
                   for b in ("Core FDE", "Adjacent")},
        "companies": {b: len({j["c"] for j in jobs if j["b"] == b}) for b in ("Core FDE", "Adjacent")},
        "cities": {b: dict(sorted(
            ((c, sum(1 for j in jobs if j["b"] == b and j["city"] == c))
             for c in {j["city"] for j in jobs if j["b"] == b}), key=lambda x: -x[1]))
            for b in ("Core FDE", "Adjacent")},
        "tiers": {b: dict(sorted(
            ((t, sum(1 for j in jobs if j["b"] == b and j["ct"] == t))
             for t in {j["ct"] for j in jobs if j["b"] == b}), key=lambda x: -x[1]))
            for b in ("Core FDE", "Adjacent")},
        "jobs": jobs,
    }
    with open(a.out, "w") as f:
        json.dump(out, f, separators=(",", ":"))

    print(f"raw    : {raw}")
    print(f"unique : {uniq}  (total {len(jobs)})")
    print(f"roles  : {roles}  (repeat postings collapsed)")
    for b in ("Core FDE", "Adjacent"):
        print(f"{b:10} levels: {out['levels'][b]}")
    print(f"written: {a.out}")

if __name__ == "__main__":
    main()
